"""Unit tests for the exam-schedule parser (Excel + PDF loaders)."""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import Workbook

from config.settings import Settings
from services import exam_schedule_parser as parser
from services.exam_schedule_parser import (
    extract_cohort,
    load_workbook_rows,
    parse_exam_workbook,
    parse_kip_time_map,
    parse_student_count,
)

HEADERS = [
    "Mã lớp QT", "Mã HP", "Tên học phần", "Ghi chú", "Nhóm", "Tuần thi",
    "Thứ", "Ngày", "Kíp thi", "Phòng thi", "SL", "Đợt", "Mã lớp thi",
]
BANNER = (
    "LỊCH THI GIỮA HỌC KỲ 2025.2 VÀ CUỐI HỌC KỲ 2025.2A\n"
    "Giờ thi: Kíp 1 (7h00) - Kíp 2 (9h30) - Kíp 3 (12h30) - Kíp 4 (15h00)"
)


def _row(**over):
    base = {
        "Mã lớp QT": "CK166692", "Mã HP": "CH1012", "Tên học phần": "Hóa học 1",
        "Ghi chú": "Vật liệu", "Nhóm": "02,04-K70C", "Tuần thi": "Tuần 35",
        "Thứ": "Thứ bảy", "Ngày": "9/5/2026", "Kíp thi": "Kíp 1",
        "Phòng thi": "D3-201", "SL": 15, "Đợt": "AB", "Mã lớp thi": "200985",
    }
    base.update(over)
    return [base[h] for h in HEADERS]


@pytest.fixture
def settings() -> Settings:
    return Settings()


# ── Excel loader ──────────────────────────────────────────────────────────────


def _write_xlsx(path, rows, *, banner=True):
    wb = Workbook()
    ws = wb.active
    if banner:
        ws.append(["LỊCH THI HỌC KỲ 2 NĂM HỌC 2025-2026"])
        ws.append([])
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_xlsx_happy_path(tmp_path, settings) -> None:
    path = tmp_path / "exam.xlsx"
    _write_xlsx(path, [_row()])
    records, report = parse_exam_workbook(str(path), settings, source_file="exam.xlsx")
    assert report.total_rows == 1 and report.valid_rows == 1
    rec = records[0]
    assert rec.subject_code == "CH1012"
    assert rec.exam_date == datetime(2026, 5, 9)
    assert rec.exam_room == "D3-201"
    assert rec.cohort == "K70C"


def test_xlsx_missing_subject_skipped(tmp_path, settings) -> None:
    path = tmp_path / "exam.xlsx"
    _write_xlsx(path, [_row(**{"Mã HP": None})])
    records, report = parse_exam_workbook(str(path), settings, source_file="exam.xlsx")
    assert records == [] and report.skipped_rows[0].reason == "no_subject"


def test_xlsx_invalid_date_skipped(tmp_path, settings) -> None:
    path = tmp_path / "exam.xlsx"
    _write_xlsx(path, [_row(**{"Ngày": "chưa có"})])
    records, report = parse_exam_workbook(str(path), settings, source_file="exam.xlsx")
    assert records == [] and report.skipped_rows[0].reason == "invalid_date"


def test_xlsx_forward_fill_date_not_subject(tmp_path, settings) -> None:
    r2 = _row(
        **{"Mã HP": "MI1141", "Tên học phần": "Giải tích 1", "Ngày": None,
           "Tuần thi": None, "Thứ": None, "Kíp thi": None, "Đợt": None,
           "Phòng thi": "D5-101"}
    )
    path = tmp_path / "exam.xlsx"
    _write_xlsx(path, [_row(), r2])
    records, report = parse_exam_workbook(str(path), settings, source_file="exam.xlsx")
    assert report.valid_rows == 2
    assert records[1].subject_code == "MI1141"  # not forward-filled
    assert records[1].exam_date == datetime(2026, 5, 9)  # forward-filled


def test_xlsx_no_header_raises(tmp_path, settings) -> None:
    wb = Workbook()
    wb.active.append(["foo", "bar", "baz"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(ValueError):
        parse_exam_workbook(str(path), settings, source_file="bad.xlsx")


# ── PDF loader (pdfplumber monkeypatched) ──────────────────────────────────────


class _FakePage:
    def __init__(self, tables, text=""):
        self._tables = tables
        self._text = text

    def extract_text(self):
        return self._text

    def extract_tables(self):
        return self._tables


class _FakePDF:
    def __init__(self, pages):
        self.pages = pages

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


def _pdf_header_row():
    # Note the PDF artifact: the last header cell carries a trailing "CK".
    return ["Mã lớp QT", "Mã HP", "Tên học phần", "Ghi chú", "Nhóm", "Tuần thi",
            "Thứ", "Ngày", "Kíp thi", "Phòng thi", "SL", "Đợt", "Mã lớp thi CK"]


def _pdf_data_row():
    return ["166692", "CH1012", "Hóa học 1", "Vật liệu 02,04-K70C", "TC",
            "Tuần 35", "Thứ bảy", "9/5/2026", "Kíp 1", "D3-201", "15", "AB", "200985"]


def _install_fake_pdf(monkeypatch, pages):
    import pdfplumber

    monkeypatch.setattr(pdfplumber, "open", lambda path: _FakePDF(pages))


def test_pdf_parses_rows_and_kip_times(monkeypatch, settings) -> None:
    banner_row = [BANNER] + [None] * 12
    table = [banner_row, _pdf_header_row(), _pdf_data_row()]
    _install_fake_pdf(monkeypatch, [_FakePage([table], text=BANNER)])

    records, report = parse_exam_workbook(
        "exam.pdf", settings, source_file="exam.pdf"
    )
    assert report.valid_rows == 1
    rec = records[0]
    assert rec.subject_code == "CH1012"
    assert rec.exam_date == datetime(2026, 5, 9)
    assert rec.exam_room == "D3-201"
    assert rec.exam_class_code == "200985"  # prefix-matched "Mã lớp thi CK"
    assert rec.cohort == "K70C"  # from the Ghi chú note
    assert rec.start_time == "07:00"  # parsed from the banner legend


def test_pdf_multipage_continuation(monkeypatch, settings) -> None:
    p1 = _FakePage([[_pdf_header_row(), _pdf_data_row()]], text=BANNER)
    # Second page: a continuation table WITHOUT a repeated header.
    row2 = _pdf_data_row()
    row2[1] = "MI1141"
    row2[9] = "D5-101"
    p2 = _FakePage([[row2]], text="")
    _install_fake_pdf(monkeypatch, [p1, p2])

    records, _ = parse_exam_workbook("exam.pdf", settings, source_file="exam.pdf")
    assert [r.subject_code for r in records] == ["CH1012", "MI1141"]
    assert records[1].exam_room == "D5-101"


def test_pdf_no_header_raises(monkeypatch, settings) -> None:
    _install_fake_pdf(monkeypatch, [_FakePage([[["a", "b", "c"]]], text="")])
    with pytest.raises(ValueError):
        parse_exam_workbook("bad.pdf", settings, source_file="bad.pdf")


# ── small helpers ───────────────────────────────────────────────────────────────


def test_parse_kip_time_map() -> None:
    m = parse_kip_time_map(BANNER)
    assert m == {"1": "07:00", "2": "09:30", "3": "12:30", "4": "15:00"}


def test_extract_cohort_scans_multiple_cells() -> None:
    assert extract_cohort("TC", "Vật liệu 02,04-K70C") == "K70C"
    assert extract_cohort("TC", "no cohort here") is None


def test_parse_student_count() -> None:
    assert parse_student_count(15) == 15
    assert parse_student_count("15 sv") == 15
    assert parse_student_count("") is None


def test_unsupported_extension_raises(settings) -> None:
    with pytest.raises(ValueError):
        parser._load_rows("schedule.csv", settings.exam_schedule_column_map)


def test_load_workbook_rows_direct(tmp_path, settings) -> None:
    path = tmp_path / "exam.xlsx"
    _write_xlsx(path, [_row()])
    rows = load_workbook_rows(str(path), settings.exam_schedule_column_map)
    assert rows and rows[0]["subject_code"] == "CH1012"
