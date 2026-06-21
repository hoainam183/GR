"""Parse a HUST exam-schedule file (``.pdf`` or ``.xlsx``) into records.

The institution distributes the schedule as a **text-based PDF**; an Excel
export with the same columns is also supported. Both resolve to the same fixed
13-column HUST schema, so a single column-name → field mapping (no LLM) and one
normalisation/validation path serve both. The only format-specific part is the
loader: ``load_pdf_rows`` (pdfplumber tables) vs ``load_workbook_rows``
(openpyxl). ``parse_exam_workbook`` dispatches by file extension and is the
coordinator. It is blocking — call ``parse_exam_workbook_async`` from async code.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import anyio

from models.exam_schedule import ExamScheduleRecord
from query.signals import fold_vietnamese_text
from schemas.exam_schedule import ParseReport, SkippedRow
from utils.vn_datetime import normalize_exam_date, normalize_session

logger = logging.getLogger(__name__)

# Canonical fields whose source cells are commonly merged across rows; a blank
# cell inherits the value above it. NEVER forward-fill identity columns
# (subject_code/exam_class_code) — that would fabricate rows.
_FORWARD_FILL_FIELDS = frozenset(
    {"exam_date", "exam_week", "exam_session", "exam_batch", "weekday"}
)
_MIN_HEADER_MATCHES = 4
_COHORT_RE = re.compile(r"K\d{2,3}[A-Za-z]?")
_INT_RE = re.compile(r"\d+")
# Banner line, e.g. "Kíp 1 (7h00) - Kíp 2 (9h30) - Kíp 3 (12h30) - Kíp 4 (15h00)".
_KIP_BANNER_RE = re.compile(r"kip\s*(\d+)\s*\(\s*(\d{1,2})\s*h\s*(\d{0,2})")

# Folded header → canonical field. The schedule ships in two column-naming
# layouts that mean the same thing; both are aliased here so a single normalised
# schema serves both (a column absent from a file simply stays at its default).
# Layout B: "Mã lớp QT / Mã HP / Tuần thi / Ngày / SL / Mã lớp thi (CK)".
# Layout A: "Mã lớp / Mã học phần / Tuần / Ngày thi / Số lượng / Mã lớp thi".
_DEFAULT_COLUMN_MAP: dict[str, str] = {
    "ma lop qt": "mgmt_class_code",
    "ma lop": "mgmt_class_code",  # layout A
    "ma hp": "subject_code",
    "ma hoc phan": "subject_code",  # layout A
    "ten hoc phan": "subject_name",
    "ghi chu": "note",
    "nhom": "group",
    "tuan thi": "exam_week",
    "tuan": "exam_week",  # layout A
    "thu": "weekday",
    "ngay": "exam_date",  # also prefix-matches layout A's "Ngày thi"
    "kip thi": "exam_session",
    "phong thi": "exam_room",
    "sl": "student_count",
    "so luong": "student_count",  # layout A
    "dot": "exam_batch",
    "ma lop thi": "exam_class_code",
}


def _fold(value: Any) -> str:
    """Fold a cell to an accent/case-insensitive, whitespace-collapsed key."""
    text = str(value if value is not None else "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", fold_vietnamese_text(text)).strip()


def _clean(value: Any) -> str:
    """Trim a cell to a display string (handles None / nbsp)."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _match_header_field(folded: str, column_map: dict[str, str]) -> str | None:
    """Map a folded header cell to a field, tolerating trailing noise.

    PDF extraction can append stray text to a header cell (e.g. the ``CK``
    suffix of the next column bleeds into ``"Mã lớp thi"`` → ``"ma lop thi ck"``).
    So accept an exact match, or a ``key + " "`` prefix match. When several keys
    prefix-match (e.g. both ``ma lop`` and ``ma lop thi`` match ``ma lop thi ck``)
    the longest key wins, so the most specific column is chosen.
    """
    if folded in column_map:
        return column_map[folded]
    best_key: str | None = None
    for key in column_map:
        if not folded.startswith(key + " "):
            continue
        if best_key is None or len(key) > len(best_key):
            best_key = key
    return column_map[best_key] if best_key is not None else None


def _find_header(
    rows: list[Any],
    column_map: dict[str, str],
) -> tuple[int, dict[int, str]] | None:
    """Locate the header row and return ``(row_index, {col_index: field})``.

    The header is the first row whose folded cells match at least
    ``_MIN_HEADER_MATCHES`` known headers — this skips title banners above it.

    Raises ``ValueError`` if such a header row is found but has no subject-code
    column: without it every data row is dropped as ``no_subject`` and the
    upload would silently store nothing. Failing loudly surfaces a 400 instead.
    """
    for row_index, cells in enumerate(rows):
        col_field: dict[int, str] = {}
        for col_index, cell in enumerate(cells):
            field = _match_header_field(_fold(cell), column_map)
            if field is not None and col_index not in col_field:
                col_field[col_index] = field
        if len(col_field) >= _MIN_HEADER_MATCHES:
            if "subject_code" not in col_field.values():
                raise ValueError(
                    "Exam-schedule header found but no subject-code column "
                    "(expected 'Mã HP' / 'Mã học phần')"
                )
            return row_index, col_field
    return None


def map_row(cells: list[Any], col_field: dict[int, str]) -> dict[str, Any]:
    """Map one source row to a canonical-field → raw-value dict."""
    fields: dict[str, Any] = {}
    for col_index, field in col_field.items():
        if col_index < len(cells):
            fields[field] = cells[col_index]
    return fields


def _mapped_data_rows(
    data_rows: list[Any],
    col_field: dict[int, str],
) -> list[dict[str, Any]]:
    """Map data rows to field dicts, dropping fully-blank / footer rows."""
    mapped: list[dict[str, Any]] = []
    for cells in data_rows:
        fields = map_row(list(cells), col_field)
        if not any(_clean(v) for v in fields.values()):
            continue
        mapped.append(fields)
    return mapped


def _apply_forward_fill(mapped: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Fill merged-cell gaps in date/week/session/batch/weekday columns."""
    last_seen: dict[str, Any] = {}
    for fields in mapped:
        for field in _FORWARD_FILL_FIELDS:
            if _clean(fields.get(field)):
                last_seen[field] = fields[field]
            elif field in last_seen:
                fields[field] = last_seen[field]
    return mapped


def load_workbook_rows(
    path: str,
    column_map: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Read an ``.xlsx`` and return mapped data rows (merged cells forward-filled).

    Raises ``ValueError`` when no recognisable header row is found.
    """
    from openpyxl import load_workbook

    column_map = column_map or _DEFAULT_COLUMN_MAP
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        raw_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    located = _find_header(raw_rows, column_map)
    if located is None:
        raise ValueError("No exam-schedule header row found in workbook")
    header_index, col_field = located
    mapped = _mapped_data_rows(raw_rows[header_index + 1 :], col_field)
    return _apply_forward_fill(mapped)


def parse_kip_time_map(text: str) -> dict[str, str]:
    """Parse the banner's "Kíp N (HhMM)" legend into ``{"N": "HH:MM"}``."""
    result: dict[str, str] = {}
    for num, hour, minute in _KIP_BANNER_RE.findall(_fold(text)):
        result[num] = f"{int(hour):02d}:{int(minute or 0):02d}"
    return result


def load_pdf_rows(
    path: str,
    column_map: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Read a text-based ``.pdf`` and return ``(mapped_rows, kip_time_map)``.

    Tables are extracted with pdfplumber. The banner + header repeat on every
    page; the header is re-detected per page and the rows below it are mapped,
    so continuation pages stay aligned. The Kíp→time legend in the banner is
    parsed for display start times. Raises ``ValueError`` when no header is found.
    """
    import pdfplumber

    column_map = column_map or _DEFAULT_COLUMN_MAP
    mapped: list[dict[str, Any]] = []
    kip_time_map: dict[str, str] = {}
    last_col_field: dict[int, str] | None = None

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            if not kip_time_map:
                kip_time_map = parse_kip_time_map(page.extract_text() or "")
            for table in page.extract_tables():
                if not table:
                    continue
                located = _find_header(table, column_map)
                if located is not None:
                    header_index, col_field = located
                    last_col_field = col_field
                    data_rows = table[header_index + 1 :]
                elif last_col_field is not None:
                    # Continuation table without a repeated header.
                    col_field = last_col_field
                    data_rows = table
                else:
                    continue
                mapped.extend(_mapped_data_rows(data_rows, col_field))

    if last_col_field is None:
        raise ValueError("No exam-schedule header row found in PDF")
    return _apply_forward_fill(mapped), kip_time_map


def extract_cohort(*values: Any) -> str | None:
    """Pull a cohort code (e.g. ``K70C``) out of any of the given cells."""
    for value in values:
        match = _COHORT_RE.search(_clean(value))
        if match:
            return match.group(0).upper()
    return None


def parse_student_count(sl_raw: Any) -> int | None:
    """Extract an integer student count from the SL cell, or ``None``."""
    if isinstance(sl_raw, bool):
        return None
    if isinstance(sl_raw, (int, float)):
        return int(sl_raw)
    match = _INT_RE.search(_clean(sl_raw))
    return int(match.group(0)) if match else None


def validate_row(fields: dict[str, Any]) -> str | None:
    """Return a skip reason, or ``None`` when the row is storable.

    Requires a subject code (Mã HP) and a parseable exam date.
    """
    if not fields.get("subject_code"):
        return "no_subject"
    if fields.get("exam_date") is None:
        return "invalid_date"
    return None


def _normalize_fields(
    raw_fields: dict[str, Any],
    *,
    date_formats: list[str],
    two_digit_year_pivot: int,
    kip_time_map: dict[str, str],
) -> dict[str, Any]:
    """Turn raw cell values into the canonical, typed field dict."""
    exam_date, exam_date_str = normalize_exam_date(
        raw_fields.get("exam_date"),
        date_formats=date_formats,
        two_digit_year_pivot=two_digit_year_pivot,
    )
    exam_session, start_time = normalize_session(
        raw_fields.get("exam_session"), kip_time_map=kip_time_map
    )
    group = _clean(raw_fields.get("group"))
    note = _clean(raw_fields.get("note"))
    return {
        "subject_code": _clean(raw_fields.get("subject_code")).upper(),
        "subject_name": _clean(raw_fields.get("subject_name")),
        "mgmt_class_code": _clean(raw_fields.get("mgmt_class_code")),
        "exam_class_code": _clean(raw_fields.get("exam_class_code")),
        "note": note,
        "group": group,
        "cohort": extract_cohort(group, note),
        "exam_week": _clean(raw_fields.get("exam_week")),
        "weekday": _clean(raw_fields.get("weekday")),
        "exam_date": exam_date,
        "exam_date_str": exam_date_str or "",
        "exam_session": exam_session or "",
        "start_time": start_time,
        "exam_room": _clean(raw_fields.get("exam_room")),
        "student_count": parse_student_count(raw_fields.get("student_count")),
        "exam_batch": _clean(raw_fields.get("exam_batch")),
    }


def _settings_value(settings: Any, name: str, default: Any) -> Any:
    value = getattr(settings, name, None)
    return default if value is None else value


def _load_rows(
    path: str,
    column_map: dict[str, str],
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Dispatch to the right loader by extension; return (rows, kip_time_map)."""
    suffix = Path(path).suffix.lower()
    if suffix == ".pdf":
        return load_pdf_rows(path, column_map)
    if suffix in (".xlsx", ".xlsm"):
        return load_workbook_rows(path, column_map), {}
    raise ValueError(f"Unsupported exam-schedule file type: {suffix!r}")


def parse_exam_workbook(
    path: str,
    settings: Any,
    *,
    source_file: str,
    source_doc_id: str | None = None,
    uploaded_by: str | None = None,
) -> tuple[list[ExamScheduleRecord], ParseReport]:
    """Parse a PDF/Excel file into records + a ParseReport. Blocking."""
    column_map = _settings_value(
        settings, "exam_schedule_column_map", _DEFAULT_COLUMN_MAP
    )
    date_formats = _settings_value(settings, "exam_schedule_date_formats", None)
    pivot = _settings_value(settings, "exam_schedule_two_digit_year_pivot", 2000)
    settings_kip = _settings_value(settings, "exam_schedule_kip_time_map", {})

    raw_rows, detected_kip = _load_rows(path, column_map)
    # The PDF banner is authoritative for this file; settings fills any gaps.
    kip_time_map = {**settings_kip, **detected_kip}

    records: list[ExamScheduleRecord] = []
    skipped: list[SkippedRow] = []
    for row_index, raw_fields in enumerate(raw_rows):
        fields = _normalize_fields(
            raw_fields,
            date_formats=date_formats,
            two_digit_year_pivot=pivot,
            kip_time_map=kip_time_map,
        )
        reason = validate_row(fields)
        if reason is not None:
            skipped.append(SkippedRow(row_index=row_index, reason=reason))
            continue
        records.append(
            ExamScheduleRecord.from_parsed_row(
                fields,
                row_index=row_index,
                source_file=source_file,
                source_doc_id=source_doc_id,
                uploaded_by=uploaded_by,
                raw={k: _clean(v) for k, v in raw_fields.items()},
            )
        )

    report = ParseReport(
        total_rows=len(raw_rows),
        valid_rows=len(records),
        skipped_rows=skipped,
    )
    logger.info(
        "Parsed exam file '%s': %d/%d valid, %d skipped",
        source_file,
        report.valid_rows,
        report.total_rows,
        len(skipped),
    )
    return records, report


async def parse_exam_workbook_async(
    path: str,
    settings: Any,
    *,
    source_file: str,
    source_doc_id: str | None = None,
    uploaded_by: str | None = None,
) -> tuple[list[ExamScheduleRecord], ParseReport]:
    """Async wrapper — offloads the blocking parse to a worker thread."""

    def _run() -> tuple[list[ExamScheduleRecord], ParseReport]:
        return parse_exam_workbook(
            path,
            settings,
            source_file=source_file,
            source_doc_id=source_doc_id,
            uploaded_by=uploaded_by,
        )

    return await anyio.to_thread.run_sync(_run)


__all__ = [
    "parse_exam_workbook",
    "parse_exam_workbook_async",
    "load_workbook_rows",
    "load_pdf_rows",
    "parse_kip_time_map",
    "map_row",
    "extract_cohort",
    "parse_student_count",
    "validate_row",
]
